import sys  # retained for print_versions()
import os
import warnings
from datetime import datetime
from typing import List, Dict, Tuple, Any

import numpy as np
import pandas as pd
import scipy
from scipy import stats
import statsmodels
import statsmodels.formula.api as smf
from statsmodels.stats.multitest import multipletests
from statsmodels.tools.sm_exceptions import PerfectSeparationError, ConvergenceWarning
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def print_versions():
    print("Python:", sys.version.replace("\n", " "))
    print("pandas:", pd.__version__)
    print("numpy:", np.__version__)
    print("scipy:", scipy.__version__)
    print("openpyxl:", openpyxl.__version__)
    print("statsmodels:", statsmodels.__version__)


def ensure_dir(path: str):
    if not os.path.isdir(path):
        os.makedirs(path, exist_ok=True)


def map_gender(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s in ("0", "0.0", "Female", "female", "F", "f"):
        return "Female"
    if s in ("1", "1.0", "Male", "male", "M", "m"):
        return "Male"
    return None


def map_binary(val, name: str):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s in ("0", "0.0"):
        return 0
    if s in ("1", "1.0"):
        return 1
    return None


def clean_categorical(series: pd.Series, allowed: List[str], var_name: str, keep_missing_as: str = None):
    s = series.copy()
    # handle missing
    if keep_missing_as is not None:
        s = s.astype(object)
        s = s.where(~s.isna(), keep_missing_as)
        s = s.astype(str)
        s = s.replace({"nan": keep_missing_as, "NaN": keep_missing_as, "": keep_missing_as})
    else:
        s = s.astype(object)
        s = s.where(~s.isna(), np.nan)
        s = s.astype(str)
        s = s.replace({"nan": np.nan, "NaN": np.nan, "": np.nan})

    # normalize numeric-like strings to int strings if possible
    def norm_val(x):
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return np.nan
        try:
            f = float(x)
            if f.is_integer():
                return str(int(f))
            return str(f)
        except Exception:
            return str(x).strip()

    s = s.map(norm_val)

    invalid_mask = ~s.isna() & ~s.isin(allowed)
    n_invalid = int(invalid_mask.sum())
    if n_invalid > 0:
        warnings.warn(f"{var_name}: {n_invalid} values not in allowed levels; set to NA")
        s.loc[invalid_mask] = np.nan
    return s


def fmt_p(p):
    if p is None or pd.isna(p):
        return "NA"
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def need_mwu(x0: pd.Series, x1: pd.Series):
    """Match baseline Table 1 rule for when to use Mann-Whitney U."""
    return (
        len(x0) < 8 or len(x1) < 8 or
        x0.var(ddof=1) == 0 or x1.var(ddof=1) == 0 or
        x0.nunique() < 3 or x1.nunique() < 3
    )


def safe_logit_fit(formula: str, data: pd.DataFrame):
    try:
        model = smf.logit(formula=formula, data=data)
    except PerfectSeparationError as e:
        return None, f"Perfect separation: {e}"
    except np.linalg.LinAlgError as e:
        return None, f"LinAlgError: {e}"
    except Exception as e:
        return None, f"Model build error: {e}"

    # Stabilized fitting attempts:
    # keep conservative reporting rule -> if still non-converged, return failure and mark as NA upstream.
    attempts = [
        {"method": "newton", "maxiter": 400},
        {"method": "lbfgs", "maxiter": 1000},
        {"method": "bfgs", "maxiter": 1000},
    ]
    fail_msgs = []

    for cfg in attempts:
        method = cfg["method"]
        maxiter = cfg["maxiter"]
        try:
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always", ConvergenceWarning)
                res = model.fit(method=method, maxiter=maxiter, disp=0)

            converged = True
            if hasattr(res, "mle_retvals") and isinstance(res.mle_retvals, dict):
                converged = bool(res.mle_retvals.get("converged", True))

            has_conv_warning = any(issubclass(w.category, ConvergenceWarning) for w in caught)

            if converged and not has_conv_warning:
                return res, None

            reasons = []
            if not converged:
                reasons.append("converged=False")
            if has_conv_warning:
                reasons.append("ConvergenceWarning")
            fail_msgs.append(f"{method}(maxiter={maxiter}): {', '.join(reasons) if reasons else 'unknown convergence issue'}")

        except PerfectSeparationError as e:
            return None, f"Perfect separation: {e}"
        except np.linalg.LinAlgError as e:
            fail_msgs.append(f"{method}(maxiter={maxiter}): LinAlgError: {e}")
        except Exception as e:
            fail_msgs.append(f"{method}(maxiter={maxiter}): Fit error: {e}")

    return None, "Non-convergence after stabilized attempts: " + " | ".join(fail_msgs)


def lr_test(full_res, reduced_res):
    try:
        lr_stat = 2.0 * (full_res.llf - reduced_res.llf)
        df_diff = int(full_res.df_model - reduced_res.df_model)
        if df_diff <= 0 or lr_stat < 0:
            return None, None, None
        p_val = stats.chi2.sf(lr_stat, df_diff)
        return lr_stat, df_diff, p_val
    except Exception:
        return None, None, None


def interaction_test_continuous(df: pd.DataFrame, var: str, subset_label: str = "All"):
    d = df[[var, "Gender", "Reabsorption"]].dropna()
    row = {
        "Variable": var,
        "Type": "Continuous",
        "Subset": subset_label,
        "N_used": len(d),
        "LR_statistic": None,
        "df_diff": None,
        "P_interaction": None,
        "Q_value_FDR_BH": None,
        "Note": "",
        "_p_raw": None,
    }
    if len(d) == 0 or d[var].nunique() < 2 or d["Gender"].nunique() < 2 or d["Reabsorption"].nunique() < 2:
        row["Note"] = "Insufficient variation or sample size"
        return row

    f_full = f"Reabsorption ~ {var} + C(Gender) + {var}:C(Gender)"
    f_red = f"Reabsorption ~ {var} + C(Gender)"
    full_res, err_full = safe_logit_fit(f_full, d)
    if full_res is None:
        row["Note"] = err_full
        return row
    red_res, err_red = safe_logit_fit(f_red, d)
    if red_res is None:
        row["Note"] = err_red
        return row

    lr_stat, df_diff, p_val = lr_test(full_res, red_res)
    if p_val is None:
        row["Note"] = "LR test unavailable (non-identifiable or numerical boundary issue)"
    row["LR_statistic"] = lr_stat
    row["df_diff"] = df_diff
    row["P_interaction"] = fmt_p(p_val)
    row["_p_raw"] = p_val
    return row


def interaction_test_categorical(df: pd.DataFrame, var: str, levels: List[str], var_type: str,
                                 var_label: str = None, subset_label: str = "All"):
    d = df[[var, "Gender", "Reabsorption"]].dropna()
    if var_label is None:
        var_label = var
    row = {
        "Variable": var_label,
        "_var_key": var,          # original column name — used as reliable key for interaction_map
        "Type": var_type,
        "Subset": subset_label,
        "N_used": len(d),
        "LR_statistic": None,
        "df_diff": None,
        "P_interaction": None,
        "Q_value_FDR_BH": None,
        "Note": "",
        "_p_raw": None,
    }
    if len(d) == 0 or d[var].nunique() < 2 or d["Gender"].nunique() < 2 or d["Reabsorption"].nunique() < 2:
        row["Note"] = "Insufficient variation or sample size"
        return row

    d = d.copy()
    d[var] = pd.Categorical(d[var], categories=levels)
    f_full = f"Reabsorption ~ C({var}) + C(Gender) + C({var}):C(Gender)"
    f_red = f"Reabsorption ~ C({var}) + C(Gender)"
    full_res, err_full = safe_logit_fit(f_full, d)
    if full_res is None:
        row["Note"] = err_full
        return row
    red_res, err_red = safe_logit_fit(f_red, d)
    if red_res is None:
        row["Note"] = err_red
        return row

    lr_stat, df_diff, p_val = lr_test(full_res, red_res)
    if p_val is None:
        row["Note"] = "LR test unavailable (non-identifiable or numerical boundary issue)"
    row["LR_statistic"] = lr_stat
    row["df_diff"] = df_diff
    row["P_interaction"] = fmt_p(p_val)
    row["_p_raw"] = p_val
    return row


def fmt_n_pct(n, denom):
    if denom is None or denom <= 0:
        return "NA"
    pct = (n / denom) * 100
    return f"{n} ({pct:.1f}%)"


def compute_chi2(table: np.ndarray):
    try:
        if table.shape[0] != 2 or table.shape[1] < 1:
            return None, None, None, None
        if table.sum() == 0:
            return None, None, None, None
        # Keep consistent with baseline script behavior (SciPy default correction handling).
        chi2, p, dof, expected = stats.chi2_contingency(table)
        return chi2, p, dof, expected
    except Exception:
        return None, None, None, None


def monte_carlo_pvalue_by_permutation(d: pd.DataFrame, var: str, levels: List[str], n_sim: int, rng: np.random.Generator):
    """
    Deprecated for primary analyses. Kept only for backward compatibility.
    """
    # observed chi2
    obs_table = np.zeros((2, len(levels)), dtype=int)
    for i, lvl in enumerate(levels):
        obs_table[0, i] = int(((d["Reabsorption"] == 0) & (d[var] == lvl)).sum())
        obs_table[1, i] = int(((d["Reabsorption"] == 1) & (d[var] == lvl)).sum())
    chi2_obs, _, dof_obs, _ = compute_chi2(obs_table)
    if chi2_obs is None:
        return None, None, None

    y = d["Reabsorption"].values
    x = d[var].values
    # precompute indices per level
    level_idx = [np.where(x == lvl)[0] for lvl in levels]
    if y.size == 0:
        return None, None, None

    chi2_sim = np.zeros(n_sim, dtype=float)
    for s in range(n_sim):
        y_perm = rng.permutation(y)
        table = np.zeros((2, len(levels)), dtype=int)
        for i, idx in enumerate(level_idx):
            if idx.size == 0:
                continue
            y_sub = y_perm[idx]
            table[0, i] = int((y_sub == 0).sum())
            table[1, i] = int((y_sub == 1).sum())
        chi2_s, _, _, _ = compute_chi2(table)
        chi2_sim[s] = 0.0 if chi2_s is None else chi2_s

    p_mc = (np.sum(chi2_sim >= chi2_obs) + 1) / (n_sim + 1)
    return chi2_obs, dof_obs, p_mc


def build_categorical_table(df: pd.DataFrame, var: str, levels: List[str], strata: List[str],
                            rng: np.random.Generator, n_sim: int = 10000):
    rows_out = []
    long_rows = []
    for stratum in strata:
        if stratum == "Total":
            d = df
        else:
            d = df[df["Gender"] == stratum]

        # contingency
        counts = {
            0: [0] * len(levels),
            1: [0] * len(levels),
        }
        for i, lvl in enumerate(levels):
            for grp in [0, 1]:
                n = int(((d["Reabsorption"] == grp) & (d[var] == lvl)).sum())
                counts[grp][i] = n

        row_totals = {grp: sum(counts[grp]) for grp in [0, 1]}
        table = np.array([counts[0], counts[1]])
        chi2, p_chi2, dof, expected = compute_chi2(table)

        test_used = "NA"
        p_report = None
        p_fisher = None
        p_mc = None

        if expected is None or chi2 is None:
            test_used = "NA"
            p_report = None
        else:
            k = table.shape[1]
            # Keep consistent with baseline script:
            # - 2x2: Fisher when any expected < 5; else Chi-square
            # - 2xk: always Chi-square; annotate when expected < 5
            low_expected = np.any(expected < 5)
            if k == 2:
                if low_expected:
                    test_used = "Fisher's exact (expected<5)"
                    try:
                        _, p_fisher = stats.fisher_exact(table)
                        p_report = p_fisher
                    except Exception:
                        test_used = "Chi-square (fallback)"
                        p_report = p_chi2
                else:
                    test_used = "Chi-square"
                    p_report = p_chi2
            else:
                if low_expected:
                    test_used = "Chi-square (expected<5; Fisher not applicable)"
                    p_report = p_chi2
                else:
                    test_used = "Chi-square"
                    p_report = p_chi2

        # Display rows for this stratum
        # Reabsorption=0
        row0 = ["Reabsorption=0"]
        for i, lvl in enumerate(levels):
            row0.append(fmt_n_pct(counts[0][i], row_totals[0]))
        row0.append("NA" if chi2 is None else f"{chi2:.3f}")
        row0.append("NA" if dof is None else f"{int(dof)}")
        row0.append(fmt_p(p_report))
        row0.append(test_used)
        # Reabsorption=1
        row1 = ["Reabsorption=1"]
        for i, lvl in enumerate(levels):
            row1.append(fmt_n_pct(counts[1][i], row_totals[1]))
        row1.append("")
        row1.append("")
        row1.append("")
        row1.append("")
        # Total row
        rowt = ["Total"]
        for i, lvl in enumerate(levels):
            rowt.append(str(counts[0][i] + counts[1][i]))
        rowt.append("")
        rowt.append("")
        rowt.append("")
        rowt.append("")

        rows_out.append((stratum, row0, row1, rowt))

        # Long format
        for grp in [0, 1]:
            for i, lvl in enumerate(levels):
                long_rows.append({
                    "variable": var,
                    "stratum": stratum,
                    "reabsorption": grp,
                    "level": lvl,
                    "n": counts[grp][i],
                    "row_total": row_totals[grp],
                    "pct": (counts[grp][i] / row_totals[grp] * 100) if row_totals[grp] > 0 else np.nan,
                    "test_used": test_used,
                    "chi2_statistic": chi2,
                    "dof": dof,
                    "p_value_reported": p_report,
                    "p_value_chi2": p_chi2,
                    "p_value_fisher": p_fisher,
                    "p_value_mc": p_mc,
                })
        for i, lvl in enumerate(levels):
            long_rows.append({
                "variable": var,
                "stratum": stratum,
                "reabsorption": "Total",
                "level": lvl,
                "n": counts[0][i] + counts[1][i],
                "row_total": counts[0][i] + counts[1][i],
                "pct": np.nan,
                "test_used": test_used,
                "chi2_statistic": chi2,
                "dof": dof,
                "p_value_reported": p_report,
                "p_value_chi2": p_chi2,
                "p_value_fisher": p_fisher,
                "p_value_mc": p_mc,
            })

    return rows_out, long_rows


def build_continuous_table(df: pd.DataFrame, var: str, strata: List[str]):
    rows_out = []
    long_rows = []
    for stratum in strata:
        d = df[df["Gender"] == stratum]
        x0 = pd.to_numeric(d.loc[d["Reabsorption"] == 0, var], errors="coerce").dropna()
        x1 = pd.to_numeric(d.loc[d["Reabsorption"] == 1, var], errors="coerce").dropna()
        n0, n1 = len(x0), len(x1)
        mean0 = np.mean(x0) if n0 > 0 else np.nan
        sd0 = np.std(x0, ddof=1) if n0 > 1 else np.nan
        mean1 = np.mean(x1) if n1 > 0 else np.nan
        sd1 = np.std(x1, ddof=1) if n1 > 1 else np.nan

        tval = pval = np.nan
        test_type = "NA"
        if n0 >= 2 and n1 >= 2:
            try:
                if need_mwu(x0, x1):
                    try:
                        _, pval = stats.mannwhitneyu(x0, x1, alternative="two-sided", method="auto")
                    except TypeError:
                        _, pval = stats.mannwhitneyu(x0, x1, alternative="two-sided")
                    tval = np.nan
                    test_type = "Mann-Whitney U"
                else:
                    tval, pval = stats.ttest_ind(x0, x1, equal_var=False)
                    test_type = "Welch"
            except Exception:
                tval, pval, test_type = np.nan, np.nan, "NA"

        row = {
            "variable": var,
            "mean_sd_0": "NA" if np.isnan(mean0) or np.isnan(sd0) else f"{mean0:.2f} ± {sd0:.2f}",
            "mean_sd_1": "NA" if np.isnan(mean1) or np.isnan(sd1) else f"{mean1:.2f} ± {sd1:.2f}",
            "t": "NA" if pd.isna(tval) else f"{tval:.3f}",
            "p": fmt_p(pval),
            "test": test_type,
            "stratum": stratum,
        }
        rows_out.append(row)

        long_rows.append({
            "variable": var,
            "stratum": stratum,
            "group": 0,
            "n": n0,
            "mean": mean0,
            "sd": sd0,
            "t": tval,
            "p": pval,
            "test": test_type,
        })
        long_rows.append({
            "variable": var,
            "stratum": stratum,
            "group": 1,
            "n": n1,
            "mean": mean1,
            "sd": sd1,
            "t": tval,
            "p": pval,
            "test": test_type,
        })

    return rows_out, long_rows


def apply_sheet_formatting(ws, max_col, header_rows: List[int], title_rows: List[int]):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # apply borders and alignment
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.border = border
                cell.alignment = center

    for r in header_rows:
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = bold
            cell.alignment = center

    for r in title_rows:
        cell = ws.cell(row=r, column=1)
        cell.font = bold
        cell.alignment = center

    # first column left align except header/title
    for r in range(1, ws.max_row + 1):
        if r in header_rows or r in title_rows:
            continue
        ws.cell(row=r, column=1).alignment = left

    # set column widths
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


def write_categorical_excel(outfile: str, cat_tables: Dict[str, Any]):
    wb = Workbook()
    wb.remove(wb.active)

    for var, (levels, rows_out) in cat_tables.items():
        ws = wb.create_sheet(title=var)
        max_col = 1 + len(levels) + 4
        header_rows = []
        title_rows = []
        row_idx = 1

        for stratum, row0, row1, rowt in rows_out:
            # Title row
            ws.cell(row=row_idx, column=1, value=stratum)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
            title_rows.append(row_idx)
            row_idx += 1

            # Header row
            headers = ["Group"] + levels + ["Chi2", "dof", "p", "Test"]
            for c, h in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=c, value=h)
            header_rows.append(row_idx)
            row_idx += 1

            # Data rows
            for row in (row0, row1, rowt):
                for c, v in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=c, value=v)
                row_idx += 1

            row_idx += 1  # spacer

        apply_sheet_formatting(ws, max_col, header_rows, title_rows)

    wb.save(outfile)


def write_continuous_excel(outfile: str, cont_rows: List[Dict[str, Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Table3_Continuous"

    max_col = 6
    header_rows = []
    title_rows = []
    row_idx = 1

    for stratum in ["Male", "Female"]:
        # Title row
        ws.cell(row=row_idx, column=1, value=stratum)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
        title_rows.append(row_idx)
        row_idx += 1

        # Header row
        headers = ["Variable", "Reabsorption=0 (Mean ± SD)", "Reabsorption=1 (Mean ± SD)", "t", "p", "Test"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=c, value=h)
        header_rows.append(row_idx)
        row_idx += 1

        # Rows for this stratum
        for r in cont_rows:
            if r["stratum"] != stratum:
                continue
            ws.cell(row=row_idx, column=1, value=r["variable"])
            ws.cell(row=row_idx, column=2, value=r["mean_sd_0"])
            ws.cell(row=row_idx, column=3, value=r["mean_sd_1"])
            ws.cell(row=row_idx, column=4, value=r["t"])
            ws.cell(row=row_idx, column=5, value=r["p"])
            ws.cell(row=row_idx, column=6, value=r["test"])
            row_idx += 1

        row_idx += 1  # spacer

    apply_sheet_formatting(ws, max_col, header_rows, title_rows)
    wb.save(outfile)

def write_interaction_excel(outfile: str, interaction_df: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "Interaction_Summary"

    headers = [
        "Variable", "Type", "Subset", "N_used",
        "LR_statistic", "df_diff", "P_interaction",
        "Q_value_FDR_BH", "Note"
    ]
    ws.append(headers)
    for _, r in interaction_df.iterrows():
        ws.append([
            r.get("Variable"),
            r.get("Type"),
            r.get("Subset"),
            r.get("N_used"),
            r.get("LR_statistic"),
            r.get("df_diff"),
            r.get("P_interaction"),
            r.get("Q_value_FDR_BH"),
            r.get("Note"),
        ])

    # formatting
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = center

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = left
        ws.cell(row=r, column=9).alignment = left

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    wb.save(outfile)


def main():
    print_versions()

    # -----------------------------------------------------------------------
    # Input file path — auto-discovery (same pattern as 5_lmm_vas_odi_joa_analysis.R)
    # Directory layout assumed:
    #   F:\李子航毕业论文原始数据\代码\   <- this script lives here
    #   F:\李子航毕业论文原始数据\文件\   <- data files live here (sibling folder)
    #
    # Candidate paths tried in order:
    #   1. <script_dir>/../文件/Retrospective data.xlsx  (standard layout)
    #   2. <working_dir>/文件/Retrospective data.xlsx     (run from project root)
    #   3. Retrospective data.xlsx                        (data in working dir, fallback)
    # -----------------------------------------------------------------------
    _target = "Retrospective data.xlsx"
    _script_path = os.path.abspath(__file__) if "__file__" in dir() else None
    _script_dir  = os.path.dirname(_script_path) if _script_path else os.getcwd()
    _candidates  = [
        os.path.normpath(os.path.join(_script_dir, "..", "文件", _target)),
        os.path.normpath(os.path.join(os.getcwd(), "文件", _target)),
        os.path.normpath(os.path.join(os.getcwd(), _target)),
    ]
    _found = [p for p in _candidates if os.path.isfile(p)]
    if not _found:
        raise FileNotFoundError(
            f"Data file '{_target}' not found. Paths tried:\n" +
            "\n".join(f"  {p}" for p in _candidates)
        )
    in_file = _found[0]
    print(f"[INFO] Retrospective data: {in_file} | sheet=Train")

    df = pd.read_excel(in_file, sheet_name="Train")

    # map gender
    df["Gender"] = df["Gender"].apply(map_gender)
    invalid_gender = df["Gender"].isna().sum()
    if invalid_gender > 0:
        warnings.warn(f"Gender: {int(invalid_gender)} invalid values removed")
    df = df[df["Gender"].isin(["Male", "Female"])].copy()

    # map outcome
    df["Reabsorption"] = df["Reabsorption"].apply(lambda x: map_binary(x, "Reabsorption"))
    invalid_outcome = df["Reabsorption"].isna().sum()
    if invalid_outcome > 0:
        warnings.warn(f"Reabsorption: {int(invalid_outcome)} invalid values removed")
    df = df[df["Reabsorption"].isin([0, 1])].copy()

    rng = np.random.default_rng(2025)

    # variable definitions
    ordinal_vars = {
        "Pfirrmann": ["1", "2", "3", "4", "5"],
        "Komori": ["1", "2", "3", "4"],
        "MSU": ["1", "2", "3"],
    }
    nominal_vars = {
        "Herniated_Level": ["L1/2", "L2/3", "L3/4", "L4/5", "L5/S1"],
        "Iwabuchi": ["1", "2", "3", "4", "5"],
        "Modic": ["0", "1", "2", "3"],
        "Spinal_canal_stenosis": ["0", "1"],
        "Bull_eye": ["1", "2", "3"],
    }

    continuous_vars = [
        "Age", "SS", "Upper_VB_Posterior_Height_CM", "Lower_VB_Posterior_Height_CM",
        "Initial_volume", "RSI", "DHI", "Months_of_Review"
    ]

    # clean categorical variables
    for var, levels in ordinal_vars.items():
        if var not in df.columns:
            df[var] = np.nan
        df[var] = clean_categorical(df[var], levels, var_name=var, keep_missing_as=None)

    for var, levels in nominal_vars.items():
        if var not in df.columns:
            df[var] = np.nan
        df[var] = clean_categorical(df[var], levels, var_name=var, keep_missing_as=None)

    # Bull_eye availability flag (0=missing, 1=1/2/3)
    df["Bull_eye_available"] = np.where(
        df["Bull_eye"].isin(["1", "2", "3"]), "1",
        np.where(df["Bull_eye"].isna(), "0", None)
    ).astype(object)

    # build categorical tables
    strata_cat = ["Male", "Female", "Total"]
    cat_tables = {}
    long_cat_rows = []
    for var, levels in {**ordinal_vars, **nominal_vars}.items():
        rows_out, long_rows = build_categorical_table(df, var, levels, strata_cat, rng=rng, n_sim=10000)
        cat_tables[var] = (levels, rows_out)
        long_cat_rows.extend(long_rows)

    # sensitivity analysis 1: Bull_eye among available scans only (1/2/3)
    df_be_avail = df[df["Bull_eye"].isin(["1", "2", "3"])].copy()
    bull_eye_avail_levels = ["1", "2", "3"]
    be_avail_rows, _ = build_categorical_table(df_be_avail, "Bull_eye", bull_eye_avail_levels, strata_cat, rng=rng, n_sim=10000)
    be_avail_tables = {"Bull_eye": (bull_eye_avail_levels, be_avail_rows)}

    # sensitivity analysis 2: Bull_eye_available flag (0/1)
    be_flag_levels = ["0", "1"]
    df["Bull_eye_available"] = clean_categorical(df["Bull_eye_available"], be_flag_levels, var_name="Bull_eye_available", keep_missing_as=None)
    be_flag_rows, _ = build_categorical_table(df, "Bull_eye_available", be_flag_levels, strata_cat, rng=rng, n_sim=10000)
    be_flag_tables = {"Bull_eye_available": (be_flag_levels, be_flag_rows)}

    # build continuous tables
    strata_cont = ["Male", "Female"]
    cont_rows = []
    long_cont_rows = []
    for var in continuous_vars:
        if var not in df.columns:
            df[var] = np.nan
        rows_out, long_rows = build_continuous_table(df, var, strata_cont)
        cont_rows.extend(rows_out)
        long_cont_rows.extend(long_rows)

    # interaction tests
    interaction_rows = []
    for var in continuous_vars:
        interaction_rows.append(interaction_test_continuous(df, var, subset_label="All"))
    for var, levels in ordinal_vars.items():
        interaction_rows.append(interaction_test_categorical(df, var, levels, "Ordinal", subset_label="All"))
    for var, levels in nominal_vars.items():
        interaction_rows.append(interaction_test_categorical(df, var, levels, "Nominal", subset_label="All"))

    interaction_df = pd.DataFrame(interaction_rows)
    interaction_df["P_interaction"] = interaction_df["P_interaction"].fillna("NA")
    valid_mask = interaction_df["_p_raw"].notna()
    if valid_mask.any():
        pvals = interaction_df.loc[valid_mask, "_p_raw"].values
        _, qvals, _, _ = multipletests(pvals, alpha=0.05, method="fdr_bh")
        interaction_df.loc[valid_mask, "Q_value_FDR_BH"] = [fmt_p(q) for q in qvals]
    interaction_df.loc[~valid_mask, "Q_value_FDR_BH"] = "NA"

    # Build interaction_map keyed on _var_key (original column name) so that
    # cat_long["variable"] / cont_long["variable"] — which always hold the raw
    # column name — can reliably look up the correct p / q values even when
    # var_label differs from var.
    interaction_map = {
        r["_var_key"]: (r.get("P_interaction"), r.get("Q_value_FDR_BH"))
        for _, r in interaction_df.iterrows()
        if pd.notna(r.get("_var_key"))
    }

    # output
    results_root = os.path.join(os.getcwd(), "Results")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_root = os.path.join(results_root, "Manuscript_v2", f"run_{run_id}")
    table2_root = os.path.join(run_root, "03_Table2_SexStratified")
    main_dir = os.path.join(table2_root, "main")
    sensitivity_dir = os.path.join(table2_root, "sensitivity")
    interaction_dir = os.path.join(table2_root, "interaction")
    audit_dir = os.path.join(table2_root, "audit")

    for d in [main_dir, sensitivity_dir, interaction_dir, audit_dir]:
        ensure_dir(d)

    cat_xlsx = os.path.join(main_dir, "Table2_Categorical.xlsx")
    cont_xlsx = os.path.join(main_dir, "Table2_Continuous.xlsx")
    write_categorical_excel(cat_xlsx, cat_tables)
    write_continuous_excel(cont_xlsx, cont_rows)

    # sensitivity outputs
    be_avail_xlsx = os.path.join(sensitivity_dir, "Table2_BullEye_available_only.xlsx")
    be_flag_xlsx = os.path.join(sensitivity_dir, "Table2_BullEye_available_flag.xlsx")
    write_categorical_excel(be_avail_xlsx, be_avail_tables)
    write_categorical_excel(be_flag_xlsx, be_flag_tables)

    # long format audit files — merge interaction p/q keyed on original column name
    cat_long = pd.DataFrame(long_cat_rows)
    cont_long = pd.DataFrame(long_cont_rows)

    cat_long[["P_interaction", "Q_value_FDR_BH"]] = cat_long["variable"].apply(
        lambda v: pd.Series(interaction_map.get(v, (np.nan, np.nan)))
    )
    cont_long[["P_interaction", "Q_value_FDR_BH"]] = cont_long["variable"].apply(
        lambda v: pd.Series(interaction_map.get(v, (np.nan, np.nan)))
    )
    cat_long.to_csv(os.path.join(audit_dir, "Table2_Categorical_long.csv"), index=False)
    cont_long.to_csv(os.path.join(audit_dir, "Table2_Continuous_long.csv"), index=False)

    # interaction outputs — drop internal keys before writing
    interaction_out_xlsx = os.path.join(interaction_dir, "Table2_Interaction_Pvalues.xlsx")
    interaction_out_csv = os.path.join(interaction_dir, "Table2_Interaction_Pvalues.csv")
    interaction_df_out = interaction_df.drop(columns=["_p_raw", "_var_key"], errors="ignore")
    write_interaction_excel(interaction_out_xlsx, interaction_df_out)
    interaction_df_out.to_csv(interaction_out_csv, index=False)

    print("[OK] Table 2 outputs written to:", table2_root)


if __name__ == "__main__":
    main()

