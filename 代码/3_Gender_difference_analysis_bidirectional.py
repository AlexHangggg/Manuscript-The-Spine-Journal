import sys
import os
import warnings
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from scipy import stats
import scipy
import statsmodels
import statsmodels.formula.api as smf
from statsmodels.stats.multitest import multipletests
from statsmodels.tools.sm_exceptions import PerfectSeparationError, ConvergenceWarning
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


def print_versions():
    print("Python:", sys.version.replace("\n", " "))
    print("pandas:", pd.__version__)
    print("numpy:", np.__version__)
    print("scipy:", scipy.__version__)
    print("openpyxl:", openpyxl.__version__)
    print("statsmodels:", statsmodels.__version__)


def ensure_dir(path: str):
    if not os.path.isdir(path):
        os.makedirs(path, exist_ok=True)


def normalize_string_value(val):
    if pd.isna(val):
        return np.nan
    s = str(val).strip()
    return np.nan if s == "" else s


def map_gender(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s in ("0", "0.0", "Female", "female", "F", "f"):
        return "Female"
    if s in ("1", "1.0", "Male", "male", "M", "m"):
        return "Male"
    return None


def map_binary(val, name: str):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s in ("0", "0.0"):
        return 0
    if s in ("1", "1.0"):
        return 1
    warnings.warn(f"{name}: invalid binary value {s!r}; set to NA")
    return None


def clean_categorical(series: pd.Series, allowed: List[str], var_name: str, keep_missing_as: Optional[str] = None):
    s = series.copy().astype(object)
    if keep_missing_as is not None:
        s = s.where(~pd.isna(s), keep_missing_as)
    else:
        s = s.where(~pd.isna(s), np.nan)

    def norm_val(x):
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return keep_missing_as if keep_missing_as is not None else np.nan
        try:
            f = float(x)
            if f.is_integer():
                return str(int(f))
            return str(f)
        except Exception:
            return str(x).strip()

    s = s.map(norm_val)
    if keep_missing_as is None:
        s = s.replace({"nan": np.nan, "NaN": np.nan, "": np.nan})
    else:
        s = s.replace({"nan": keep_missing_as, "NaN": keep_missing_as, "": keep_missing_as})

    invalid_mask = ~pd.isna(s) & ~s.isin(allowed)
    n_invalid = int(invalid_mask.sum())
    if n_invalid > 0:
        warnings.warn(f"{var_name}: {n_invalid} values not in allowed levels; set to NA")
        s.loc[invalid_mask] = np.nan if keep_missing_as is None else keep_missing_as
    return s


def fmt_p(p):
    if p is None or pd.isna(p):
        return "NA"
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def need_mwu(x0: pd.Series, x1: pd.Series):
    return (
        len(x0) < 8 or len(x1) < 8 or
        x0.var(ddof=1) == 0 or x1.var(ddof=1) == 0 or
        x0.nunique() < 3 or x1.nunique() < 3
    )


def safe_logit_fit(formula: str, data: pd.DataFrame):
    try:
        model = smf.logit(formula=formula, data=data)
    except PerfectSeparationError as e:
        return None, f"Perfect separation: {e}"
    except np.linalg.LinAlgError as e:
        return None, f"LinAlgError: {e}"
    except Exception as e:
        return None, f"Model build error: {e}"

    attempts = [
        {"method": "newton", "maxiter": 400},
        {"method": "lbfgs", "maxiter": 1000},
        {"method": "bfgs", "maxiter": 1000},
    ]
    fail_msgs = []
    for cfg in attempts:
        try:
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always", ConvergenceWarning)
                res = model.fit(method=cfg["method"], maxiter=cfg["maxiter"], disp=0)
            converged = True
            if hasattr(res, "mle_retvals") and isinstance(res.mle_retvals, dict):
                converged = bool(res.mle_retvals.get("converged", True))
            has_warning = any(issubclass(w.category, ConvergenceWarning) for w in caught)
            if converged and not has_warning:
                return res, None
            reasons = []
            if not converged:
                reasons.append("converged=False")
            if has_warning:
                reasons.append("ConvergenceWarning")
            fail_msgs.append(f"{cfg['method']}: {', '.join(reasons) if reasons else 'fit issue'}")
        except PerfectSeparationError as e:
            return None, f"Perfect separation: {e}"
        except np.linalg.LinAlgError as e:
            fail_msgs.append(f"{cfg['method']}: LinAlgError: {e}")
        except Exception as e:
            fail_msgs.append(f"{cfg['method']}: Fit error: {e}")
    return None, "Non-convergence after stabilized attempts: " + " | ".join(fail_msgs)


def lr_test(full_res, reduced_res):
    try:
        lr_stat = 2.0 * (full_res.llf - reduced_res.llf)
        df_diff = int(full_res.df_model - reduced_res.df_model)
        if df_diff <= 0 or lr_stat < 0:
            return None, None, None
        return lr_stat, df_diff, stats.chi2.sf(lr_stat, df_diff)
    except Exception:
        return None, None, None


def interaction_test_continuous(df: pd.DataFrame, var: str, subset_label: str = "All", adjust_cohort: bool = False):
    cols = [var, "Gender", "Reabsorption"]
    if adjust_cohort and "Cohort" in df.columns and var != "Cohort":
        cols.append("Cohort")
    d = df[cols].dropna()
    row = {
        "Variable": var,
        "_var_key": var,
        "Type": "Continuous",
        "Subset": subset_label,
        "N_used": len(d),
        "LR_statistic": None,
        "df_diff": None,
        "P_interaction": None,
        "Q_value_FDR_BH": None,
        "Note": "",
        "_p_raw": None,
    }
    if len(d) == 0 or d[var].nunique() < 2 or d["Gender"].nunique() < 2 or d["Reabsorption"].nunique() < 2:
        row["Note"] = "Insufficient variation or sample size"
        return row

    if adjust_cohort and "Cohort" in d.columns and var != "Cohort" and d["Cohort"].nunique() >= 2:
        f_full = f"Reabsorption ~ {var} + C(Gender) + C(Cohort) + {var}:C(Gender)"
        f_red = f"Reabsorption ~ {var} + C(Gender) + C(Cohort)"
    else:
        f_full = f"Reabsorption ~ {var} + C(Gender) + {var}:C(Gender)"
        f_red = f"Reabsorption ~ {var} + C(Gender)"
    full_res, err_full = safe_logit_fit(f_full, d)
    if full_res is None:
        row["Note"] = err_full
        return row
    red_res, err_red = safe_logit_fit(f_red, d)
    if red_res is None:
        row["Note"] = err_red
        return row

    lr_stat, df_diff, p_val = lr_test(full_res, red_res)
    if p_val is None:
        row["Note"] = "LR test unavailable"
    row["LR_statistic"] = lr_stat
    row["df_diff"] = df_diff
    row["P_interaction"] = fmt_p(p_val)
    row["_p_raw"] = p_val
    return row


def interaction_test_categorical(df: pd.DataFrame, var: str, levels: List[str], var_type: str,
                                 subset_label: str = "All", adjust_cohort: bool = False):
    cols = [var, "Gender", "Reabsorption"]
    if adjust_cohort and "Cohort" in df.columns and var != "Cohort":
        cols.append("Cohort")
    d = df[cols].dropna()
    row = {
        "Variable": var,
        "_var_key": var,
        "Type": var_type,
        "Subset": subset_label,
        "N_used": len(d),
        "LR_statistic": None,
        "df_diff": None,
        "P_interaction": None,
        "Q_value_FDR_BH": None,
        "Note": "",
        "_p_raw": None,
    }
    if len(d) == 0 or d[var].nunique() < 2 or d["Gender"].nunique() < 2 or d["Reabsorption"].nunique() < 2:
        row["Note"] = "Insufficient variation or sample size"
        return row

    d = d.copy()
    d[var] = pd.Categorical(d[var], categories=levels)
    if adjust_cohort and "Cohort" in d.columns and var != "Cohort" and d["Cohort"].nunique() >= 2:
        f_full = f"Reabsorption ~ C({var}) + C(Gender) + C(Cohort) + C({var}):C(Gender)"
        f_red = f"Reabsorption ~ C({var}) + C(Gender) + C(Cohort)"
    else:
        f_full = f"Reabsorption ~ C({var}) + C(Gender) + C({var}):C(Gender)"
        f_red = f"Reabsorption ~ C({var}) + C(Gender)"
    full_res, err_full = safe_logit_fit(f_full, d)
    if full_res is None:
        row["Note"] = err_full
        return row
    red_res, err_red = safe_logit_fit(f_red, d)
    if red_res is None:
        row["Note"] = err_red
        return row

    lr_stat, df_diff, p_val = lr_test(full_res, red_res)
    if p_val is None:
        row["Note"] = "LR test unavailable"
    row["LR_statistic"] = lr_stat
    row["df_diff"] = df_diff
    row["P_interaction"] = fmt_p(p_val)
    row["_p_raw"] = p_val
    return row


def fmt_n_pct(n, denom):
    if denom is None or denom <= 0:
        return "NA"
    pct = (n / denom) * 100
    return f"{n} ({pct:.1f}%)"


def detect_categorical_pvalue(d: pd.DataFrame, var: str, levels: List[str]):
    """Return (p_raw, p_fmt, test_name, stat) where stat is Chi2 value or None."""
    d = d[[var, "Reabsorption"]].dropna()
    if len(d) == 0 or d["Reabsorption"].nunique() < 2 or d[var].nunique() < 2:
        return None, "NA", "Insufficient data", None
    d = d.copy()
    d[var] = pd.Categorical(d[var], categories=levels)
    table = pd.crosstab(d["Reabsorption"], d[var], dropna=False)
    if table.shape[0] < 2:
        return None, "NA", "Insufficient data", None
    observed = table.to_numpy()
    try:
        chi2, p, dof, expected = stats.chi2_contingency(observed)
        if observed.shape == (2, 2) and np.any(expected < 5):
            p = stats.fisher_exact(observed)[1]
            # Fisher's exact does not produce a Chi2 statistic
            return p, fmt_p(p), "Fisher's exact", None
        return p, fmt_p(p), "Chi-square", chi2
    except Exception:
        return None, "NA", "Test failed", None


def detect_continuous_pvalue(x0: pd.Series, x1: pd.Series):
    if len(x0) == 0 or len(x1) == 0:
        return None, "NA", "Insufficient data", None
    try:
        if need_mwu(x0, x1):
            stat, p = stats.mannwhitneyu(x0, x1, alternative="two-sided")
            return stat, fmt_p(p), "Mann-Whitney U", p
        stat, p = stats.ttest_ind(x0, x1, equal_var=False, nan_policy="omit")
        return stat, fmt_p(p), "Welch t-test", p
    except Exception:
        return None, "NA", "Test failed", None


def build_categorical_table(df: pd.DataFrame, var: str, levels: List[str], strata: List[str]):
    rows_out = []
    long_rows = []
    for stratum in strata:
        d = df if stratum == "Total" else df[df["Gender"] == stratum]
        p_raw, p_fmt, test_name, stat = detect_categorical_pvalue(d, var, levels)
        stat_fmt = "" if stat is None or pd.isna(stat) else f"{stat:.3f}"
        rows_out.append({
            "section": var,
            "label": stratum,
            "reabs0": "",
            "reabs1": "",
            "stat": stat_fmt,
            "p": p_fmt,
            "test": test_name,
        })
        for lvl in levels:
            n0 = int(((d["Reabsorption"] == 0) & (d[var] == lvl)).sum())
            n1 = int(((d["Reabsorption"] == 1) & (d[var] == lvl)).sum())
            d0 = int((d["Reabsorption"] == 0).sum())
            d1 = int((d["Reabsorption"] == 1).sum())
            rows_out.append({
                "section": var,
                "label": f"  {lvl}",
                "reabs0": fmt_n_pct(n0, d0),
                "reabs1": fmt_n_pct(n1, d1),
                "stat": "",
                "p": "",
                "test": "",
            })
            long_rows.append({
                "variable": var,
                "stratum": stratum,
                "level": lvl,
                "n_reabs0": n0,
                "n_reabs1": n1,
                "denom_reabs0": d0,
                "denom_reabs1": d1,
                "pct_reabs0": (n0 / d0 * 100) if d0 > 0 else np.nan,
                "pct_reabs1": (n1 / d1 * 100) if d1 > 0 else np.nan,
                "test": test_name,
                "statistic": stat,
                "p_raw": p_raw,
                "p": p_fmt,
            })
    return rows_out, long_rows


def build_continuous_table(df: pd.DataFrame, var: str, strata: List[str]):
    rows_out = []
    long_rows = []
    for stratum in strata:
        d = df[df["Gender"] == stratum]
        x0 = pd.to_numeric(d.loc[d["Reabsorption"] == 0, var], errors="coerce").dropna()
        x1 = pd.to_numeric(d.loc[d["Reabsorption"] == 1, var], errors="coerce").dropna()
        stat, p_fmt, test_name, p_raw = detect_continuous_pvalue(x0, x1)
        mean_sd_0 = f"{x0.mean():.2f} ± {x0.std(ddof=1):.2f}" if len(x0) > 0 else "NA"
        mean_sd_1 = f"{x1.mean():.2f} ± {x1.std(ddof=1):.2f}" if len(x1) > 0 else "NA"
        rows_out.append({
            "variable": var,
            "stratum": stratum,
            "mean_sd_0": mean_sd_0,
            "mean_sd_1": mean_sd_1,
            "t": "" if stat is None or pd.isna(stat) else f"{stat:.3f}",
            "p": p_fmt,
            "test": test_name,
        })
        long_rows.append({
            "variable": var,
            "stratum": stratum,
            "n_reabs0": len(x0),
            "n_reabs1": len(x1),
            "mean_reabs0": x0.mean() if len(x0) > 0 else np.nan,
            "sd_reabs0": x0.std(ddof=1) if len(x0) > 1 else np.nan,
            "mean_reabs1": x1.mean() if len(x1) > 0 else np.nan,
            "sd_reabs1": x1.std(ddof=1) if len(x1) > 1 else np.nan,
            "statistic": stat,
            "p_raw": p_raw,
            "p": p_fmt,
            "test": test_name,
        })
    return rows_out, long_rows


def apply_sheet_formatting(ws, max_col: int):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = center
        row[0].alignment = left
    for i in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22


def write_categorical_excel(outfile: str, cat_tables: Dict[str, Any]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorical_Table2"
    headers = ["Variable", "Stratum/Level", "Reabsorption=0", "Reabsorption=1", "Statistic", "P", "Test"]
    ws.append(headers)
    for var, (_, rows) in cat_tables.items():
        for row in rows:
            ws.append([
                var if row["label"] == "Male" else "",
                row["label"],
                row["reabs0"],
                row["reabs1"],
                row["stat"],
                row["p"],
                row["test"],
            ])
    apply_sheet_formatting(ws, len(headers))
    wb.save(outfile)


def write_continuous_excel(outfile: str, cont_rows: List[Dict[str, Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Continuous_Table2"
    headers = ["Variable", "Stratum", "Reabsorption=0 (Mean ± SD)", "Reabsorption=1 (Mean ± SD)", "Statistic", "P", "Test"]
    ws.append(headers)
    for row in cont_rows:
        ws.append([
            row["variable"],
            row["stratum"],
            row["mean_sd_0"],
            row["mean_sd_1"],
            row["t"],
            row["p"],
            row["test"],
        ])
    apply_sheet_formatting(ws, len(headers))
    wb.save(outfile)


def write_interaction_excel(outfile: str, interaction_df: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "Interaction_Summary"
    headers = ["Variable", "Type", "Subset", "N_used", "LR_statistic", "df_diff", "P_interaction", "Q_value_FDR_BH", "Note"]
    ws.append(headers)
    for _, r in interaction_df.iterrows():
        ws.append([
            r.get("Variable"),
            r.get("Type"),
            r.get("Subset"),
            r.get("N_used"),
            r.get("LR_statistic"),
            r.get("df_diff"),
            r.get("P_interaction"),
            r.get("Q_value_FDR_BH"),
            r.get("Note"),
        ])
    apply_sheet_formatting(ws, len(headers))
    wb.save(outfile)


def discover_input_file(target: str) -> str:
    script_path = os.path.abspath(__file__) if "__file__" in dir() else None
    script_dir = os.path.dirname(script_path) if script_path else os.getcwd()
    candidates = [
        os.path.normpath(os.path.join(script_dir, "..", "文件", target)),
        os.path.normpath(os.path.join(os.getcwd(), "文件", target)),
        os.path.normpath(os.path.join(os.getcwd(), target)),
    ]
    found = [p for p in candidates if os.path.isfile(p)]
    if not found:
        raise FileNotFoundError(
            f"Data file '{target}' not found. Paths tried:\n" +
            "\n".join(f"  {p}" for p in candidates)
        )
    return found[0]


def normalize_age_key(val):
    if pd.isna(val):
        return None
    try:
        num = float(val)
        if np.isnan(num):
            return None
        if num.is_integer():
            return str(int(num))
        return f"{num:.6f}".rstrip("0").rstrip(".")
    except Exception:
        s = str(val).strip()
        return s if s else None


def standardize_text_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_object_dtype(out[col]) or pd.api.types.is_string_dtype(out[col]):
            out[col] = out[col].apply(normalize_string_value)
    return out


def build_overlap_key(df: pd.DataFrame) -> pd.Series:
    names = df["Name"].apply(normalize_string_value)
    ages = df["Age"].apply(normalize_age_key)
    genders = df["Gender"].apply(map_gender)
    keys = []
    for name, age, gender in zip(names, ages, genders):
        if pd.isna(name) or age is None or gender is None:
            keys.append(None)
        else:
            keys.append(f"{name}|{age}|{gender}")
    return pd.Series(keys, index=df.index, dtype="object")


def load_bidirectional_dataset():
    retro_file = discover_input_file("Retrospective data.xlsx")
    pros_file = discover_input_file("Prospective data.xlsx")

    retro = pd.read_excel(retro_file, sheet_name="Train")
    pros = pd.read_excel(pros_file, sheet_name="Train_Pors")
    if list(retro.columns) != list(pros.columns):
        retro_only = [c for c in retro.columns if c not in pros.columns]
        pros_only = [c for c in pros.columns if c not in retro.columns]
        raise ValueError(
            "Retrospective and prospective columns are not identical.\n"
            f"Retrospective-only: {retro_only}\nProspective-only: {pros_only}"
        )

    retro = standardize_text_columns(retro)
    pros = standardize_text_columns(pros)

    retro["Cohort"] = "Retrospective"
    pros["Cohort"] = "Prospective"
    retro["Source_File"] = os.path.basename(retro_file)
    pros["Source_File"] = os.path.basename(pros_file)
    retro["Source_Sheet"] = "Train"
    pros["Source_Sheet"] = "Train_Pors"
    retro["Unified_ID"] = retro["Cohort"].astype(str) + "_" + retro["ID"].astype(str)
    pros["Unified_ID"] = pros["Cohort"].astype(str) + "_" + pros["ID"].astype(str)

    retro["_overlap_key"] = build_overlap_key(retro)
    pros["_overlap_key"] = build_overlap_key(pros)
    overlap_keys = sorted(set(retro["_overlap_key"].dropna()) & set(pros["_overlap_key"].dropna()))

    retro_overlap = retro[retro["_overlap_key"].isin(overlap_keys)].copy()
    pros_overlap = pros[pros["_overlap_key"].isin(overlap_keys)].copy()

    overlap_matches = retro_overlap[["Unified_ID", "ID", "Name", "Age", "Gender", "_overlap_key"]].merge(
        pros_overlap[["Unified_ID", "ID", "Name", "Age", "Gender", "_overlap_key"]],
        on="_overlap_key",
        how="inner",
        suffixes=("_retro", "_pros"),
    ).rename(columns={"_overlap_key": "Overlap_Key"})

    # Keep same-person records across cohorts as distinct longitudinal observations.
    merged = pd.concat([retro, pros], ignore_index=True)
    merged = merged.drop(columns=["_overlap_key"], errors="ignore")
    retro_drop = retro_overlap.iloc[0:0].drop(columns=["_overlap_key"], errors="ignore")
    summary = {
        "retro_file": retro_file,
        "pros_file": pros_file,
        "retro_n": int(len(retro)),
        "pros_n": int(len(pros)),
        "overlap_key_n": int(len(overlap_keys)),
        "retro_dropped_n": 0,
        "merged_n": int(len(merged)),
        "merged_cohort_counts": merged["Cohort"].value_counts(dropna=False).to_dict(),
    }
    return merged, overlap_matches, retro_drop, summary


def write_merge_audit(base_dir: str, merged_df: pd.DataFrame, overlap_matches: pd.DataFrame,
                      retro_drop: pd.DataFrame, summary: Dict[str, Any]):
    ensure_dir(base_dir)
    overlap_matches.to_csv(os.path.join(base_dir, "overlap_matches.csv"), index=False)
    retro_drop.to_csv(os.path.join(base_dir, "dropped_retrospective_due_to_overlap.csv"), index=False)
    merged_df.to_csv(os.path.join(base_dir, "merged_analysis_dataset.csv"), index=False)

    lines = [
        "Bidirectional merge summary",
        f"Retrospective file: {summary['retro_file']}",
        f"Prospective file: {summary['pros_file']}",
        f"Retrospective rows before merge: {summary['retro_n']}",
        f"Prospective rows before merge: {summary['pros_n']}",
        f"Overlap keys detected (Name+Age+Gender): {summary['overlap_key_n']}",
        "Retrospective rows dropped due to overlap: 0 (overlaps retained as distinct cohort-specific observations)",
        f"Merged rows with overlap retained: {summary['merged_n']}",
    ]
    for cohort, count in summary["merged_cohort_counts"].items():
        lines.append(f"Merged cohort count - {cohort}: {count}")
    with open(os.path.join(base_dir, "merge_summary.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def build_analysis_outputs(df: pd.DataFrame, adjust_cohort: bool, include_cohort_var: bool, subset_label: str):
    df = df.copy()
    df["Gender"] = df["Gender"].apply(map_gender)
    invalid_gender = df["Gender"].isna().sum()
    if invalid_gender > 0:
        warnings.warn(f"Gender: {int(invalid_gender)} invalid values removed ({subset_label})")
    df = df[df["Gender"].isin(["Male", "Female"])].copy()

    df["Reabsorption"] = df["Reabsorption"].apply(lambda x: map_binary(x, "Reabsorption"))
    invalid_outcome = df["Reabsorption"].isna().sum()
    if invalid_outcome > 0:
        warnings.warn(f"Reabsorption: {int(invalid_outcome)} invalid values removed ({subset_label})")
    df = df[df["Reabsorption"].isin([0, 1])].copy()

    ordinal_vars = {
        "Pfirrmann": ["1", "2", "3", "4", "5"],
        "Komori": ["1", "2", "3", "4"],
        "MSU": ["1", "2", "3"],
    }
    nominal_vars = {
        "Herniated_Level": ["L1/2", "L2/3", "L3/4", "L4/5", "L5/S1"],
        "Iwabuchi": ["1", "2", "3", "4", "5"],
        "Modic": ["0", "1", "2", "3"],
        "Spinal_canal_stenosis": ["0", "1"],
        "Bull_eye": ["1", "2", "3"],
    }
    if include_cohort_var and "Cohort" in df.columns:
        nominal_vars["Cohort"] = ["Retrospective", "Prospective"]
    continuous_vars = [
        "Age", "SS", "Upper_VB_Posterior_Height_CM", "Lower_VB_Posterior_Height_CM",
        "Initial_volume", "RSI", "DHI", "Months_of_Review",
    ]

    for var, levels in ordinal_vars.items():
        if var not in df.columns:
            df[var] = np.nan
        df[var] = clean_categorical(df[var], levels, var_name=var, keep_missing_as=None)
    for var, levels in nominal_vars.items():
        if var not in df.columns:
            df[var] = np.nan
        df[var] = clean_categorical(df[var], levels, var_name=var, keep_missing_as=None)

    df["Bull_eye_available"] = np.where(
        df["Bull_eye"].isin(["1", "2", "3"]),
        "1",
        np.where(df["Bull_eye"].isna(), "0", None),
    ).astype(object)

    strata_cat = ["Male", "Female", "Total"]
    cat_tables = {}
    long_cat_rows = []
    for var, levels in {**ordinal_vars, **nominal_vars}.items():
        rows, long_rows = build_categorical_table(df, var, levels, strata_cat)
        cat_tables[var] = (levels, rows)
        long_cat_rows.extend(long_rows)

    df_be_avail = df[df["Bull_eye"].isin(["1", "2", "3"])].copy()
    be_avail_levels = ["1", "2", "3"]
    be_avail_rows, _ = build_categorical_table(df_be_avail, "Bull_eye", be_avail_levels, strata_cat)
    be_avail_tables = {"Bull_eye": (be_avail_levels, be_avail_rows)}

    be_flag_levels = ["0", "1"]
    df["Bull_eye_available"] = clean_categorical(
        df["Bull_eye_available"], be_flag_levels, var_name="Bull_eye_available", keep_missing_as=None
    )
    be_flag_rows, _ = build_categorical_table(df, "Bull_eye_available", be_flag_levels, strata_cat)
    be_flag_tables = {"Bull_eye_available": (be_flag_levels, be_flag_rows)}

    strata_cont = ["Male", "Female"]
    cont_rows = []
    long_cont_rows = []
    for var in continuous_vars:
        if var not in df.columns:
            df[var] = np.nan
        rows, long_rows = build_continuous_table(df, var, strata_cont)
        cont_rows.extend(rows)
        long_cont_rows.extend(long_rows)

    interaction_rows = []
    for var in continuous_vars:
        interaction_rows.append(interaction_test_continuous(df, var, subset_label=subset_label, adjust_cohort=adjust_cohort))
    for var, levels in ordinal_vars.items():
        interaction_rows.append(interaction_test_categorical(df, var, levels, "Ordinal", subset_label=subset_label, adjust_cohort=adjust_cohort))
    for var, levels in nominal_vars.items():
        interaction_rows.append(interaction_test_categorical(df, var, levels, "Nominal", subset_label=subset_label, adjust_cohort=adjust_cohort))

    interaction_df = pd.DataFrame(interaction_rows)
    interaction_df["P_interaction"] = interaction_df["P_interaction"].fillna("NA")
    valid_mask = interaction_df["_p_raw"].notna()
    if valid_mask.any():
        _, qvals, _, _ = multipletests(interaction_df.loc[valid_mask, "_p_raw"].values, alpha=0.05, method="fdr_bh")
        interaction_df.loc[valid_mask, "Q_value_FDR_BH"] = [fmt_p(q) for q in qvals]
    interaction_df.loc[~valid_mask, "Q_value_FDR_BH"] = "NA"

    interaction_map = {
        r["_var_key"]: (r.get("P_interaction"), r.get("Q_value_FDR_BH"))
        for _, r in interaction_df.iterrows()
        if pd.notna(r.get("_var_key"))
    }

    cat_long = pd.DataFrame(long_cat_rows)
    if not cat_long.empty:
        cat_long[["P_interaction", "Q_value_FDR_BH"]] = cat_long["variable"].apply(
            lambda v: pd.Series(interaction_map.get(v, (np.nan, np.nan)))
        )
    cont_long = pd.DataFrame(long_cont_rows)
    if not cont_long.empty:
        cont_long[["P_interaction", "Q_value_FDR_BH"]] = cont_long["variable"].apply(
            lambda v: pd.Series(interaction_map.get(v, (np.nan, np.nan)))
        )

    return {
        "analysis_df": df,
        "cat_tables": cat_tables,
        "cont_rows": cont_rows,
        "be_avail_tables": be_avail_tables,
        "be_flag_tables": be_flag_tables,
        "cat_long": cat_long,
        "cont_long": cont_long,
        "interaction_df_out": interaction_df.drop(columns=["_p_raw", "_var_key"], errors="ignore"),
    }


def write_main_outputs(output: Dict[str, Any], table2_root: str):
    main_dir = os.path.join(table2_root, "main")
    sensitivity_dir = os.path.join(table2_root, "sensitivity")
    interaction_dir = os.path.join(table2_root, "interaction")
    audit_dir = os.path.join(table2_root, "audit")
    for d in [main_dir, sensitivity_dir, interaction_dir, audit_dir]:
        ensure_dir(d)

    write_categorical_excel(os.path.join(main_dir, "Table2_Categorical.xlsx"), output["cat_tables"])
    write_continuous_excel(os.path.join(main_dir, "Table2_Continuous.xlsx"), output["cont_rows"])
    write_categorical_excel(os.path.join(sensitivity_dir, "Table2_BullEye_available_only.xlsx"), output["be_avail_tables"])
    write_categorical_excel(os.path.join(sensitivity_dir, "Table2_BullEye_available_flag.xlsx"), output["be_flag_tables"])
    output["cat_long"].to_csv(os.path.join(audit_dir, "Table2_Categorical_long.csv"), index=False)
    output["cont_long"].to_csv(os.path.join(audit_dir, "Table2_Continuous_long.csv"), index=False)
    write_interaction_excel(os.path.join(interaction_dir, "Table2_Interaction_Pvalues.xlsx"), output["interaction_df_out"])
    output["interaction_df_out"].to_csv(os.path.join(interaction_dir, "Table2_Interaction_Pvalues.csv"), index=False)


def write_flat_outputs(output: Dict[str, Any], out_dir: str):
    ensure_dir(out_dir)
    write_categorical_excel(os.path.join(out_dir, "Table2_Categorical.xlsx"), output["cat_tables"])
    write_continuous_excel(os.path.join(out_dir, "Table2_Continuous.xlsx"), output["cont_rows"])
    write_interaction_excel(os.path.join(out_dir, "Table2_Interaction_Pvalues.xlsx"), output["interaction_df_out"])
    output["interaction_df_out"].to_csv(os.path.join(out_dir, "Table2_Interaction_Pvalues.csv"), index=False)
    output["cat_long"].to_csv(os.path.join(out_dir, "Table2_Categorical_long.csv"), index=False)
    output["cont_long"].to_csv(os.path.join(out_dir, "Table2_Continuous_long.csv"), index=False)


def main():
    print_versions()
    results_root = os.path.join(os.getcwd(), "Results")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_root = os.path.join(results_root, "Manuscript_v2", f"run_{run_id}")
    table2_root = os.path.join(run_root, "03_Table2_SexStratified_bidirectional")
    merge_audit_dir = os.path.join(table2_root, "merge_audit")

    merged_df, overlap_matches, retro_drop, summary = load_bidirectional_dataset()
    write_merge_audit(merge_audit_dir, merged_df, overlap_matches, retro_drop, summary)

    main_output = build_analysis_outputs(merged_df, adjust_cohort=True, include_cohort_var=True, subset_label="All_bidirectional")
    write_main_outputs(main_output, table2_root)

    by_cohort_root = os.path.join(table2_root, "sensitivity", "by_cohort")
    for cohort in ["Retrospective", "Prospective"]:
        cohort_df = merged_df[merged_df["Cohort"] == cohort].copy()
        cohort_output = build_analysis_outputs(cohort_df, adjust_cohort=False, include_cohort_var=False, subset_label=f"{cohort}_only")
        write_flat_outputs(cohort_output, os.path.join(by_cohort_root, cohort))

    print("[OK] Bidirectional Table 2 outputs written to:", table2_root)


if __name__ == "__main__":
    main()
